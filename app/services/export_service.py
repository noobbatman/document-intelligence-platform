"""Export service — CSV, Excel (XLSX), and JSON export of extraction results.

One of the most-requested features: every commercial tool exports to CSV/Excel.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import Document, ExtractionResult


def _flatten_fields(doc: Document, result: ExtractionResult | None) -> dict[str, Any]:
    """Flatten a document + extraction result into a single-row dict."""
    fields = result.export_payload.get("fields", {}) if result else {}
    return {
        "document_id": doc.id,
        "filename": doc.filename,
        "document_type": doc.document_type or "",
        "status": doc.status,
        "document_confidence": doc.document_confidence,
        "classifier_confidence": doc.classifier_confidence,
        "pipeline_version": doc.pipeline_version,
        "tenant_id": doc.tenant_id or "",
        "created_at": doc.created_at.isoformat() if doc.created_at else "",
        # Extracted fields
        "invoice_number": fields.get("invoice_number", ""),
        "invoice_date": fields.get("invoice_date", ""),
        "due_date": fields.get("due_date", ""),
        "vendor_name": fields.get("vendor_name", ""),
        "customer_name": fields.get("customer_name", ""),
        "subtotal": fields.get("subtotal", ""),
        "tax": fields.get("tax", ""),
        "total_amount": fields.get("total_amount", ""),
        "currency": fields.get("currency", ""),
        "account_number": fields.get("account_number", ""),
        "statement_period": fields.get("statement_period", ""),
        "opening_balance": fields.get("opening_balance", ""),
        "closing_balance": fields.get("closing_balance", ""),
        "available_balance": fields.get("available_balance", ""),
    }


class ExportService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def _query_documents(
        self,
        *,
        document_type: str | None = None,
        status: str | None = None,
        tenant_id: str | None = None,
        since: datetime | None = None,
        limit: int = 10_000,
    ) -> list[Document]:
        stmt = select(Document).where(Document.deleted_at.is_(None))
        if document_type:
            stmt = stmt.where(Document.document_type == document_type)
        if status:
            stmt = stmt.where(Document.status == status)
        if tenant_id:
            stmt = stmt.where(Document.tenant_id == tenant_id)
        if since:
            stmt = stmt.where(Document.created_at >= since)
        stmt = stmt.order_by(Document.created_at.desc()).limit(limit)
        return list(self.db.scalars(stmt))

    # ── CSV ───────────────────────────────────────────────────────────────────

    def export_csv(self, **filters) -> bytes:
        """Return UTF-8 encoded CSV bytes."""
        docs = self._query_documents(**filters)
        if not docs:
            return b""

        rows = [_flatten_fields(d, d.extraction_result) for d in docs]
        fieldnames = list(rows[0].keys()) if rows else []

        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        return buf.getvalue().encode("utf-8")

    # ── Excel (XLSX) ──────────────────────────────────────────────────────────

    def export_xlsx(self, **filters) -> bytes:
        """Return XLSX bytes using openpyxl (optional dependency)."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise RuntimeError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        docs = self._query_documents(**filters)
        rows = [_flatten_fields(d, d.extraction_result) for d in docs]
        if not rows:
            rows = [{}]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Documents"

        # Header row styling
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        fieldnames = list(rows[0].keys()) if rows else []

        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── JSON batch export ─────────────────────────────────────────────────────

    def export_json_batch(self, **filters) -> bytes:
        """Return full extraction payloads as a JSON array."""
        docs = self._query_documents(**filters)
        output = []
        for doc in docs:
            payload = doc.extraction_result.export_payload if doc.extraction_result else {}
            output.append(
                {
                    "document_id": doc.id,
                    "filename": doc.filename,
                    "document_type": doc.document_type,
                    "status": doc.status,
                    "created_at": doc.created_at.isoformat() if doc.created_at else None,
                    "export_payload": payload,
                }
            )
        return json.dumps(output, indent=2).encode("utf-8")
